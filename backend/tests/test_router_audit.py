"""Route-level tests for /api/audit-log endpoints (list, filters, export)."""

from datetime import UTC, datetime
from io import BytesIO
from unittest.mock import MagicMock
from uuid import uuid4

import httpx
from openpyxl import load_workbook

from app.exports.audit_excel_generator import generate_audit_excel

XLSX_MEDIA_TYPE = "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"


def _mock_entry(*, action="priority_created", entity_type="cve_priority", entity_id="CVE-2024-0001", user_id=None):
    e = MagicMock()
    e.id = uuid4()
    e.user_id = user_id
    e.action = action
    e.entity_type = entity_type
    e.entity_id = entity_id
    e.details = {"level": "high"}
    e.created_at = datetime(2024, 5, 1, 12, 30, 0, tzinfo=UTC)
    return e


# -- GET /api/audit-log --


async def test_list_requires_sec_team(team_member_client: httpx.AsyncClient):
    resp = await team_member_client.get("/api/audit-log")
    assert resp.status_code == 403


async def test_list_returns_paginated(sec_team_client: httpx.AsyncClient, mock_app_db):
    resp = await sec_team_client.get("/api/audit-log")
    assert resp.status_code == 200
    body = resp.json()
    assert set(body) == {"items", "total", "page", "page_size"}


async def test_list_accepts_filters(sec_team_client: httpx.AsyncClient, mock_app_db):
    resp = await sec_team_client.get(
        "/api/audit-log?search=cve&action=priority_created&entity_type=cve_priority&date_from=2024-01-01&date_to=2024-12-31"
    )
    assert resp.status_code == 200


async def test_list_invalid_date_returns_400(sec_team_client: httpx.AsyncClient, mock_app_db):
    resp = await sec_team_client.get("/api/audit-log?date_from=not-a-date")
    assert resp.status_code == 400


# -- GET /api/audit-log/filters --


async def test_filters_requires_sec_team(team_member_client: httpx.AsyncClient):
    resp = await team_member_client.get("/api/audit-log/filters")
    assert resp.status_code == 403


async def test_filters_returns_distinct_lists(sec_team_client: httpx.AsyncClient, mock_app_db):
    actions_result = MagicMock()
    actions_result.scalars.return_value.all.return_value = ["priority_created", "settings_updated"]
    entity_result = MagicMock()
    entity_result.scalars.return_value.all.return_value = ["cve_priority", "global_settings"]
    mock_app_db.execute.side_effect = [actions_result, entity_result]

    resp = await sec_team_client.get("/api/audit-log/filters")
    assert resp.status_code == 200
    body = resp.json()
    assert body["actions"] == ["priority_created", "settings_updated"]
    assert body["entity_types"] == ["cve_priority", "global_settings"]


# -- GET /api/audit-log/export --


async def test_export_requires_sec_team(team_member_client: httpx.AsyncClient):
    resp = await team_member_client.get("/api/audit-log/export")
    assert resp.status_code == 403


async def test_export_returns_xlsx_de_filename(sec_team_client: httpx.AsyncClient, mock_app_db):
    result = MagicMock()
    result.scalars.return_value.all.return_value = [_mock_entry()]
    mock_app_db.execute.return_value = result

    resp = await sec_team_client.get("/api/audit-log/export")
    assert resp.status_code == 200
    assert resp.headers["content-type"] == XLSX_MEDIA_TYPE
    assert "audit-protokoll" in resp.headers["content-disposition"]
    assert len(resp.content) > 0


async def test_export_en_filename(sec_team_client: httpx.AsyncClient, mock_app_db):
    result = MagicMock()
    result.scalars.return_value.all.return_value = []
    mock_app_db.execute.return_value = result

    resp = await sec_team_client.get("/api/audit-log/export?lang=en")
    assert resp.status_code == 200
    assert "audit-log" in resp.headers["content-disposition"]


# -- generate_audit_excel (pure) --


def test_generate_audit_excel_localizes_action():
    rows = [
        {
            "created_at": datetime(2024, 5, 1, 12, 30, 0),
            "username": "secadmin",
            "action": "priority_created",
            "entity_type": "cve_priority",
            "entity_id": "CVE-2024-0001",
            "details": {"level": "high"},
        }
    ]
    data = generate_audit_excel(rows, lang="en")
    wb = load_workbook(BytesIO(data))
    ws = wb.active
    assert ws.cell(row=1, column=1).value == "Timestamp"
    assert ws.cell(row=2, column=2).value == "secadmin"
    assert ws.cell(row=2, column=3).value == "Priority set"
    assert "level: high" in ws.cell(row=2, column=6).value


def test_generate_audit_excel_unknown_action_humanized():
    rows = [{"created_at": datetime(2024, 5, 1), "action": "some_new_action", "entity_type": "x", "details": {}}]
    data = generate_audit_excel(rows, lang="de")
    wb = load_workbook(BytesIO(data))
    assert wb.active.cell(row=2, column=3).value == "some new action"
