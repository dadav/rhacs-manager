"""Excel export/import for CVE data and risk acceptance batch creation."""

from datetime import datetime
from io import BytesIO

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

TRANSLATIONS: dict[str, dict] = {
    "de": {
        "severity_labels": {0: "Unbekannt", 1: "Gering", 2: "Mittel", 3: "Hoch", 4: "Kritisch"},
        "priority_labels": {"critical": "Kritisch", "high": "Hoch", "medium": "Mittel", "low": "Gering"},
        "ra_status_labels": {
            "requested": "Beantragt",
            "approved": "Genehmigt",
            "rejected": "Abgelehnt",
            "expired": "Abgelaufen",
        },
        "yes": "Ja",
        "no": "Nein",
        "sheet_data": "CVE-Daten",
        "sheet_help": "Anleitung",
        "help_title": "CVE-Import Anleitung",
        "columns": [
            ("CVE-ID", 18),
            ("Schweregrad", 14),
            ("CVSS", 8),
            ("EPSS", 10),
            ("Komponente", 22),
            ("Version", 14),
            ("Behebbar", 10),
            ("Fix-Version", 16),
            ("Image", 40),
            ("Erstmals gesehen", 16),
            ("Veröffentlicht", 16),
            ("Priorität", 12),
            ("RA-Status", 14),
            ("RA-Begründung (ausfüllen)", 44),
            ("RA-Ablaufdatum (optional, JJJJ-MM-TT)", 34),
        ],
        "instructions": """ANLEITUNG — CVE-Risikoakzeptanz per Excel-Import

So erstellen Sie Risikoakzeptanzen über diesen Excel-Import:

1. BEGRÜNDUNG AUSFÜLLEN
   Tragen Sie in der Spalte "RA-Begründung (ausfüllen)" (Spalte N) Ihre Begründung für die Risikoakzeptanz ein.
   - Mindestens 10 Zeichen, maximal 5000 Zeichen.
   - Nur Zeilen mit ausgefüllter Begründung werden importiert.

2. ABLAUFDATUM (OPTIONAL)
   Tragen Sie in der Spalte "RA-Ablaufdatum (optional, JJJJ-MM-TT)" (Spalte O) ein optionales Ablaufdatum ein.
   - Format: JJJJ-MM-TT (z.B. 2025-12-31)
   - Wenn leer, wird kein Ablaufdatum gesetzt.

3. GELTUNGSBEREICH
   Beim Import wird der Geltungsbereich automatisch ermittelt:
   - Alle betroffenen Namespaces des Benutzers werden für die CVE berücksichtigt (Namespace-Scope).

4. IMPORT
   Laden Sie die ausgefüllte Datei im CVE-Manager hoch:
   - Schwachstellen → "Excel importieren"
   - Vorschau prüfen → "Importieren" bestätigen

HINWEISE:
- Ändern Sie NICHT die Spaltenreihenfolge oder -namen in Sheet "CVE-Daten".
- Nur Team-Mitglieder können importieren (Security-Team hat keine Import-Berechtigung).
- Bereits existierende aktive Risikoakzeptanzen für dieselbe CVE+Scope werden nicht dupliziert.
""",
    },
    "en": {
        "severity_labels": {0: "Unknown", 1: "Low", 2: "Medium", 3: "High", 4: "Critical"},
        "priority_labels": {"critical": "Critical", "high": "High", "medium": "Medium", "low": "Low"},
        "ra_status_labels": {
            "requested": "Requested",
            "approved": "Approved",
            "rejected": "Rejected",
            "expired": "Expired",
        },
        "yes": "Yes",
        "no": "No",
        "sheet_data": "CVE Data",
        "sheet_help": "Instructions",
        "help_title": "CVE Import Instructions",
        "columns": [
            ("CVE-ID", 18),
            ("Severity", 14),
            ("CVSS", 8),
            ("EPSS", 10),
            ("Component", 22),
            ("Version", 14),
            ("Fixable", 10),
            ("Fix Version", 16),
            ("Image", 40),
            ("First Seen", 16),
            ("Published", 16),
            ("Priority", 12),
            ("RA Status", 14),
            ("RA Justification (fill in)", 44),
            ("RA Expiry Date (optional, YYYY-MM-DD)", 34),
        ],
        "instructions": """INSTRUCTIONS — CVE Risk Acceptance via Excel Import

How to create risk acceptances using this Excel import:

1. FILL IN JUSTIFICATION
   Enter your justification for the risk acceptance in column "RA Justification (fill in)" (column N).
   - Minimum 10 characters, maximum 5000 characters.
   - Only rows with a filled justification will be imported.

2. EXPIRY DATE (OPTIONAL)
   Enter an optional expiry date in column "RA Expiry Date (optional, YYYY-MM-DD)" (column O).
   - Format: YYYY-MM-DD (e.g. 2025-12-31)
   - If empty, no expiry date will be set.

3. SCOPE
   The scope is determined automatically during import:
   - All affected namespaces of the user are considered for the CVE (namespace scope).

4. IMPORT
   Upload the completed file in the CVE Manager:
   - Vulnerabilities → "Import Excel"
   - Review preview → Confirm "Import"

NOTES:
- Do NOT change the column order or names in the "CVE Data" sheet.
- Only team members can import (security team does not have import permission).
- Existing active risk acceptances for the same CVE+scope will not be duplicated.
""",
    },
}

# Column index mapping: maps both German and English column names to a canonical key.
# Used by the import parser to accept files exported in either language.
_COLUMN_CANONICAL: dict[str, str] = {}
for _lang_key in ("de", "en"):
    _cols = TRANSLATIONS[_lang_key]["columns"]
    # Canonical keys are the English column names (index 0 of each tuple)
    _en_cols = TRANSLATIONS["en"]["columns"]
    for _i, (_name, _) in enumerate(_cols):
        _COLUMN_CANONICAL[_name] = _en_cols[_i][0]

# The two editable column canonical keys
_JUSTIFICATION_KEY = "RA Justification (fill in)"
_EXPIRY_KEY = "RA Expiry Date (optional, YYYY-MM-DD)"

# All known sheet data names (for import parser)
_SHEET_DATA_NAMES = {TRANSLATIONS[lang]["sheet_data"] for lang in TRANSLATIONS}


def _get_translations(lang: str) -> dict:
    return TRANSLATIONS.get(lang, TRANSLATIONS["de"])


def generate_cve_excel(rows: list[dict], lang: str = "de") -> bytes:
    """Generate an Excel workbook with CVE data and instruction sheet.

    Each dict in `rows` should contain:
      cve_id, severity, cvss, epss_probability, component_name, component_version,
      fixable, fixed_by, image_name, first_seen, published_on, priority_level,
      risk_acceptance_status
    """
    t = _get_translations(lang)
    data_columns = t["columns"]
    severity_labels = t["severity_labels"]
    priority_labels = t["priority_labels"]
    ra_status_labels = t["ra_status_labels"]
    yes_label = t["yes"]
    no_label = t["no"]

    wb = Workbook()

    # Sheet 1: CVE data
    ws = wb.active
    ws.title = t["sheet_data"]

    header_font = Font(bold=True, color="FFFFFF", size=11)
    header_fill = PatternFill(start_color="0066CC", end_color="0066CC", fill_type="solid")

    # Find indices for editable columns (Justification, Expiry) and EPSS
    col_names = [name for name, _ in data_columns]
    justification_col_name = TRANSLATIONS[lang]["columns"][13][0] if lang in TRANSLATIONS else data_columns[13][0]
    expiry_col_name = TRANSLATIONS[lang]["columns"][14][0] if lang in TRANSLATIONS else data_columns[14][0]
    epss_col_name = data_columns[3][0]  # EPSS is always at index 3

    begr_col = col_names.index(justification_col_name) + 1
    ablauf_col = col_names.index(expiry_col_name) + 1
    epss_col = col_names.index(epss_col_name) + 1
    editable_cols = [begr_col, ablauf_col]

    for col_idx, (col_name, col_width) in enumerate(data_columns, start=1):
        cell = ws.cell(row=1, column=col_idx, value=col_name)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center", vertical="center")
        ws.column_dimensions[cell.column_letter].width = col_width

    # Highlight editable column headers
    editable_fill = PatternFill(start_color="FFF3CD", end_color="FFF3CD", fill_type="solid")
    for col_idx in editable_cols:
        cell = ws.cell(row=1, column=col_idx)
        cell.fill = PatternFill(start_color="EC7A08", end_color="EC7A08", fill_type="solid")

    for row_idx, row_data in enumerate(rows, start=2):
        severity = row_data.get("severity", 0)
        first_seen = row_data.get("first_seen")
        published_on = row_data.get("published_on")

        values = [
            row_data.get("cve_id", ""),
            severity_labels.get(severity, severity_labels.get(0, "Unknown")),
            round(row_data.get("cvss", 0), 1),
            row_data.get("epss_probability", 0),
            row_data.get("component_name", ""),
            row_data.get("component_version", ""),
            yes_label if row_data.get("fixable") else no_label,
            row_data.get("fixed_by", "") or "",
            row_data.get("image_name", ""),
            first_seen.strftime("%Y-%m-%d") if isinstance(first_seen, datetime) else str(first_seen or ""),
            published_on.strftime("%Y-%m-%d") if isinstance(published_on, datetime) else str(published_on or ""),
            priority_labels.get(row_data.get("priority_level", ""), row_data.get("priority_level", "")) or "",
            ra_status_labels.get(row_data.get("risk_acceptance_status", ""), row_data.get("risk_acceptance_status", ""))
            or "",
            "",  # Justification (empty for user to fill)
            "",  # Expiry date (empty for user to fill)
        ]

        for col_idx, val in enumerate(values, start=1):
            cell = ws.cell(row=row_idx, column=col_idx, value=val)
            if col_idx == epss_col:
                cell.number_format = "0.00%"
            if col_idx in editable_cols:
                cell.fill = editable_fill

    last_col = get_column_letter(len(data_columns))
    ws.auto_filter.ref = f"A1:{last_col}{len(rows) + 1}"
    ws.freeze_panes = "A2"

    # Sheet 2: Instructions
    ws_help = wb.create_sheet(t["sheet_help"])
    ws_help.column_dimensions["A"].width = 100
    ws_help.cell(row=1, column=1, value=t["help_title"]).font = Font(bold=True, size=14)
    for line_idx, line in enumerate(t["instructions"].strip().split("\n"), start=3):
        ws_help.cell(row=line_idx, column=1, value=line)

    buf = BytesIO()
    wb.save(buf)
    return buf.getvalue()


def parse_import_excel(file_bytes: bytes) -> list[dict]:
    """Parse an uploaded Excel file and extract rows with non-empty justification.

    Accepts files exported in either German or English (sheet name and column headers).

    Returns a list of dicts with keys:
      cve_id, justification, expires_at
    """
    wb = load_workbook(BytesIO(file_bytes), read_only=True, data_only=True)

    # Find the data sheet by trying known names
    sheet_name = None
    for candidate in _SHEET_DATA_NAMES:
        if candidate in wb.sheetnames:
            sheet_name = candidate
            break
    if sheet_name is None:
        names_str = " / ".join(sorted(_SHEET_DATA_NAMES))
        raise ValueError(f"Sheet '{names_str}' nicht gefunden / not found")

    ws = wb[sheet_name]

    # Read header row and map to canonical keys
    headers = [cell.value for cell in next(ws.iter_rows(min_row=1, max_row=1))]

    # Validate that all expected canonical columns are present
    expected_canonical = {name for name, _ in TRANSLATIONS["en"]["columns"]}
    present = {_COLUMN_CANONICAL.get(h, h) for h in headers if h is not None}
    missing = expected_canonical - present
    if missing:
        raise ValueError(f"Fehlende Spalten / Missing columns: {', '.join(sorted(missing))}")

    # Build column map using canonical keys
    col_map: dict[str, int] = {}
    for idx, name in enumerate(headers):
        if name is not None:
            canonical = _COLUMN_CANONICAL.get(name, name)
            col_map[canonical] = idx

    results: list[dict] = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        justification_idx = col_map.get(_JUSTIFICATION_KEY)
        if justification_idx is None:
            continue
        justification = row[justification_idx]
        if not justification or not str(justification).strip():
            continue

        justification = str(justification).strip()

        # Parse expires_at
        expires_at_idx = col_map.get(_EXPIRY_KEY)
        expires_at = None
        if expires_at_idx is not None:
            raw_date = row[expires_at_idx]
            if raw_date:
                raw_str = str(raw_date).strip()
                if raw_str:
                    try:
                        expires_at = datetime.strptime(raw_str[:10], "%Y-%m-%d")
                    except ValueError:
                        if isinstance(raw_date, datetime):
                            expires_at = raw_date

        cve_id_idx = col_map.get("CVE-ID")
        cve_id = str(row[cve_id_idx] or "").strip() if cve_id_idx is not None else ""

        results.append(
            {
                "cve_id": cve_id,
                "justification": justification,
                "expires_at": expires_at,
            }
        )

    wb.close()
    return results
