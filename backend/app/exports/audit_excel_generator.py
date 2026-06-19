"""Excel export for audit log entries.

Mirrors the styling of ``excel_generator.generate_cve_excel`` (blue header row,
auto-filter, frozen header) so audit exports look consistent with CVE exports.
Action labels are localized server-side from the dicts below so the downloaded
file is self-describing without the frontend. German is the default/fallback.
"""

from datetime import datetime
from io import BytesIO

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

# Localized column headers (name, width) and human-readable action labels.
# Keep both languages in sync; unknown actions fall back to a humanized string.
_TRANSLATIONS: dict[str, dict] = {
    "de": {
        "sheet": "Audit-Log",
        "columns": [
            ("Zeitstempel", 20),
            ("Benutzer", 22),
            ("Aktion", 32),
            ("Entitätstyp", 20),
            ("Entitäts-ID", 38),
            ("Details", 80),
        ],
        "actions": {
            "priority_created": "Priorität gesetzt",
            "priority_updated": "Priorität aktualisiert",
            "priority_deleted": "Priorität entfernt",
            "risk_acceptance_created": "Risikoakzeptanz beantragt",
            "risk_acceptance_auto_approved": "Risikoakzeptanz automatisch genehmigt",
            "risk_acceptance_updated": "Risikoakzeptanz aktualisiert",
            "risk_acceptance_reviewed": "Risikoakzeptanz geprüft",
            "risk_acceptance_assigned": "Risikoakzeptanz zugewiesen",
            "risk_acceptance_deleted": "Risikoakzeptanz gelöscht",
            "risk_acceptance_imported": "Risikoakzeptanz importiert",
            "risk_acceptance_expired": "Risikoakzeptanz abgelaufen",
            "remediation_created": "Behebung erstellt",
            "remediation_updated": "Behebung aktualisiert",
            "remediation_deleted": "Behebung gelöscht",
            "suppression_rule_created": "Unterdrückungsregel erstellt",
            "suppression_rule_updated": "Unterdrückungsregel aktualisiert",
            "suppression_rule_reviewed": "Unterdrückungsregel überprüft",
            "suppression_rule_deleted": "Unterdrückungsregel gelöscht",
            "settings_updated": "Einstellungen aktualisiert",
            "digest_manual_send": "Digest manuell versendet",
        },
    },
    "en": {
        "sheet": "Audit Log",
        "columns": [
            ("Timestamp", 20),
            ("User", 22),
            ("Action", 32),
            ("Entity type", 20),
            ("Entity ID", 38),
            ("Details", 80),
        ],
        "actions": {
            "priority_created": "Priority set",
            "priority_updated": "Priority updated",
            "priority_deleted": "Priority removed",
            "risk_acceptance_created": "Risk acceptance requested",
            "risk_acceptance_auto_approved": "Risk acceptance auto-approved",
            "risk_acceptance_updated": "Risk acceptance updated",
            "risk_acceptance_reviewed": "Risk acceptance reviewed",
            "risk_acceptance_assigned": "Risk acceptance assigned",
            "risk_acceptance_deleted": "Risk acceptance deleted",
            "risk_acceptance_imported": "Risk acceptance imported",
            "risk_acceptance_expired": "Risk acceptance expired",
            "remediation_created": "Remediation created",
            "remediation_updated": "Remediation updated",
            "remediation_deleted": "Remediation deleted",
            "suppression_rule_created": "Suppression rule created",
            "suppression_rule_updated": "Suppression rule updated",
            "suppression_rule_reviewed": "Suppression rule reviewed",
            "suppression_rule_deleted": "Suppression rule deleted",
            "settings_updated": "Settings updated",
            "digest_manual_send": "Digest sent manually",
        },
    },
}


def _get_translations(lang: str) -> dict:
    return _TRANSLATIONS.get(lang, _TRANSLATIONS["de"])


def _format_details(details: object) -> str:
    """Flatten the JSON details dict into a readable ``key: value; ...`` string."""
    if isinstance(details, dict):
        return "; ".join(f"{k}: {v}" for k, v in details.items())
    return str(details or "")


def generate_audit_excel(rows: list[dict], lang: str = "de") -> bytes:
    """Generate an Excel workbook with audit log entries.

    Each dict in ``rows`` should contain:
      created_at (datetime), username, action, entity_type, entity_id, details
    """
    t = _get_translations(lang)
    columns = t["columns"]
    action_labels = t["actions"]

    wb = Workbook()
    ws = wb.active
    ws.title = t["sheet"]

    header_font = Font(bold=True, color="FFFFFF", size=11)
    header_fill = PatternFill(start_color="0066CC", end_color="0066CC", fill_type="solid")

    for col_idx, (col_name, col_width) in enumerate(columns, start=1):
        cell = ws.cell(row=1, column=col_idx, value=col_name)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center", vertical="center")
        ws.column_dimensions[cell.column_letter].width = col_width

    for row_idx, row in enumerate(rows, start=2):
        created = row.get("created_at")
        timestamp = created.strftime("%Y-%m-%d %H:%M:%S") if isinstance(created, datetime) else str(created or "")
        action = row.get("action", "")
        action_label = action_labels.get(action, action.replace("_", " "))

        values = [
            timestamp,
            row.get("username") or "",
            action_label,
            row.get("entity_type", ""),
            row.get("entity_id") or "",
            _format_details(row.get("details")),
        ]
        for col_idx, val in enumerate(values, start=1):
            ws.cell(row=row_idx, column=col_idx, value=val)

    last_col = get_column_letter(len(columns))
    ws.auto_filter.ref = f"A1:{last_col}{len(rows) + 1}"
    ws.freeze_panes = "A2"

    buf = BytesIO()
    wb.save(buf)
    return buf.getvalue()
